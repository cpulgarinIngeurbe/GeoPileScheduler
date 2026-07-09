"""Script para crear archivo Excel de prueba CASO 6 - Optimización compleja."""

import openpyxl
from datetime import datetime


def crear_caso_prueba_6():
    """Crea CASO 6: 2 unidades, 4 equipos, 20 pilotes (optimización compleja)."""
    wb = openpyxl.Workbook()

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Proyecto
    ws = wb.create_sheet("Proyecto", 0)
    ws["A1:D1"] = ["Proyecto", "FechaInicio", "RadioCritico_m", "TiempoRestriccion_h"]
    ws["A2"] = "Caso 6 - Optimización Compleja"
    ws["B2"] = datetime(2026, 1, 1)
    ws["C2"] = 7.0  # Radio pequeño → muchos bloqueos
    ws["D2"] = 24.0
    ws["B2"].number_format = "YYYY-MM-DD"

    # Unidades: 2 unidades
    ws = wb.create_sheet("Unidades", 1)
    ws["A1:B1"] = ["UnidadID", "Nombre"]
    ws["A2:B2"] = ["TORRE_A", "Torre A"]
    ws["A3:B3"] = ["TORRE_B", "Torre B"]

    # Pilotes: 20 pilotes (10 por unidad en grid)
    ws = wb.create_sheet("Pilotes", 2)
    ws["A1:D1"] = ["PiloteID", "UnidadID", "X", "Y"]

    pilotes_data = []
    # Torre A: 10 pilotes en grid 5x2
    for i in range(5):
        for j in range(2):
            pilotes_data.append(
                (f"PA_{i:02d}{j:02d}", "TORRE_A", float(i * 6), float(j * 6))
            )

    # Torre B: 10 pilotes en grid 5x2 (desplazado)
    for i in range(5):
        for j in range(2):
            pilotes_data.append(
                (f"PB_{i:02d}{j:02d}", "TORRE_B", float(i * 6 + 1), float(j * 6 + 1))
            )

    for idx, (pid, uid, x, y) in enumerate(pilotes_data, start=2):
        ws[f"A{idx}"] = pid
        ws[f"B{idx}"] = uid
        ws[f"C{idx}"] = x
        ws[f"D{idx}"] = y

    # Equipos: 4 equipos con distinto rendimiento
    ws = wb.create_sheet("Equipos", 3)
    ws["A1:G1"] = ["EquipoID", "Nombre", "RendimientoPilotesDia", "ModoInicio", "PiloteInicio", "ModoFin", "PiloteFin"]

    equipos_data = [
        ("EQ_1", "Equipo 1 (Rápido)", 4, "AUTO", "", "AUTO", ""),
        ("EQ_2", "Equipo 2 (Medio)", 3, "AUTO", "", "AUTO", ""),
        ("EQ_3", "Equipo 3 (Medio)", 2, "AUTO", "", "AUTO", ""),
        ("EQ_4", "Equipo 4 (Lento)", 1, "AUTO", "", "AUTO", ""),
    ]

    for idx, (eid, nombre, rend, m_ini, p_ini, m_fin, p_fin) in enumerate(equipos_data, start=2):
        ws[f"A{idx}"] = eid
        ws[f"B{idx}"] = nombre
        ws[f"C{idx}"] = rend
        ws[f"D{idx}"] = m_ini
        ws[f"E{idx}"] = p_ini
        ws[f"F{idx}"] = m_fin
        ws[f"G{idx}"] = p_fin

    # Asignaciones: Equipos pueden trabajar en ambas unidades
    ws = wb.create_sheet("AsignacionEquipos", 4)
    ws["A1:C1"] = ["EquipoID", "UnidadID", "Prioridad"]

    asignaciones = [
        ("EQ_1", "TORRE_A", 1),
        ("EQ_1", "TORRE_B", 2),
        ("EQ_2", "TORRE_A", 3),
        ("EQ_2", "TORRE_B", 4),
        ("EQ_3", "TORRE_A", 5),
        ("EQ_3", "TORRE_B", 6),
        ("EQ_4", "TORRE_A", 7),
        ("EQ_4", "TORRE_B", 8),
    ]

    for idx, (eid, uid, prio) in enumerate(asignaciones, start=2):
        ws[f"A{idx}"] = eid
        ws[f"B{idx}"] = uid
        ws[f"C{idx}"] = prio

    # Ancho columnas
    for sheet in wb.worksheets:
        for col in sheet.columns:
            sheet.column_dimensions[col[0].column_letter].width = 20

    output_path = "data/input/caso_prueba_6.xlsx"
    wb.save(output_path)
    print(f"✅ Archivo creado: {output_path}")


if __name__ == "__main__":
    crear_caso_prueba_6()
