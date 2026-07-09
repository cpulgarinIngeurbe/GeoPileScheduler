"""Script para crear archivo Excel de prueba CASO 1.

CASO 1 especificación:
- 1 unidad estructural (Torre 1)
- 1 equipo de pilotaje (Equipo A)
- 5 pilotes en patrón circular
- Inicio MANUAL, Fin MANUAL
"""

from datetime import datetime
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


def crear_caso_prueba_1():
    """Crea archivo Excel para CASO 1."""
    # Crear workbook
    wb = openpyxl.Workbook()

    # Eliminar hoja por defecto
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # 1. Hoja Proyecto
    ws_proyecto = wb.create_sheet("Proyecto", 0)
    ws_proyecto["A1"] = "Proyecto"
    ws_proyecto["B1"] = "FechaInicio"
    ws_proyecto["C1"] = "RadioCritico_m"
    ws_proyecto["D1"] = "TiempoRestriccion_h"

    ws_proyecto["A2"] = "Caso Prueba 1 - Pilotaje Simple"
    ws_proyecto["B2"] = datetime(2026, 1, 1)
    ws_proyecto["C2"] = 10.0  # Radio de 10 metros
    ws_proyecto["D2"] = 24.0  # Tiempo mínimo 24 horas

    # Formato fecha
    ws_proyecto["B2"].number_format = "YYYY-MM-DD"

    # 2. Hoja Unidades
    ws_unidades = wb.create_sheet("Unidades", 1)
    ws_unidades["A1"] = "UnidadID"
    ws_unidades["B1"] = "Nombre"

    ws_unidades["A2"] = "TORRE_1"
    ws_unidades["B2"] = "Torre 1"

    # 3. Hoja Pilotes
    ws_pilotes = wb.create_sheet("Pilotes", 2)
    ws_pilotes["A1"] = "PiloteID"
    ws_pilotes["B1"] = "UnidadID"
    ws_pilotes["C1"] = "X"
    ws_pilotes["D1"] = "Y"

    # 5 pilotes en círculo con radio 5 metros (para estar dentro del radio crítico)
    pilotes = [
        ("P_001", "TORRE_1", 0.0, 0.0),      # Centro
        ("P_002", "TORRE_1", 5.0, 0.0),      # Este
        ("P_003", "TORRE_1", 3.54, 3.54),    # NE
        ("P_004", "TORRE_1", 0.0, 5.0),      # Norte
        ("P_005", "TORRE_1", -3.54, 3.54),   # NO
    ]

    for idx, (pid, uid, x, y) in enumerate(pilotes, start=2):
        ws_pilotes[f"A{idx}"] = pid
        ws_pilotes[f"B{idx}"] = uid
        ws_pilotes[f"C{idx}"] = x
        ws_pilotes[f"D{idx}"] = y

    # 4. Hoja Equipos
    ws_equipos = wb.create_sheet("Equipos", 3)
    ws_equipos["A1"] = "EquipoID"
    ws_equipos["B1"] = "Nombre"
    ws_equipos["C1"] = "RendimientoPilotesDia"
    ws_equipos["D1"] = "ModoInicio"
    ws_equipos["E1"] = "PiloteInicio"
    ws_equipos["F1"] = "ModoFin"
    ws_equipos["G1"] = "PiloteFin"

    ws_equipos["A2"] = "EQUIPO_A"
    ws_equipos["B2"] = "Equipo A"
    ws_equipos["C2"] = 2  # 2 pilotes por día
    ws_equipos["D2"] = "MANUAL"
    ws_equipos["E2"] = "P_001"  # Comienza en P_001
    ws_equipos["F2"] = "MANUAL"
    ws_equipos["G2"] = "P_005"  # Termina en P_005

    # 5. Hoja AsignacionEquipos
    ws_asignaciones = wb.create_sheet("AsignacionEquipos", 4)
    ws_asignaciones["A1"] = "EquipoID"
    ws_asignaciones["B1"] = "UnidadID"
    ws_asignaciones["C1"] = "Prioridad"

    ws_asignaciones["A2"] = "EQUIPO_A"
    ws_asignaciones["B2"] = "TORRE_1"
    ws_asignaciones["C2"] = 1

    # Ajustar ancho de columnas
    for ws in [ws_proyecto, ws_unidades, ws_pilotes, ws_equipos, ws_asignaciones]:
        for column in ws.columns:
            max_length = 20
            column_letter = column[0].column_letter
            ws.column_dimensions[column_letter].width = max_length

    # Guardar
    output_path = "data/input/caso_prueba_1.xlsx"
    wb.save(output_path)
    print(f"✅ Archivo creado: {output_path}")
    return output_path


if __name__ == "__main__":
    crear_caso_prueba_1()
