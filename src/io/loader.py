"""Cargador de Excel para GeoPile Scheduler.

Responsabilidad: Leer archivo Excel con estructura oficial (5 hojas)
y construir el modelo interno del proyecto.
"""

from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from src.core.models import Proyecto, UnidadEstructural, Pilote, Equipo, AsignacionEquipo
from src.core.enums import ModoInicio, ModoFin
from src.io.validator import validate_project_data


def load_project(filepath: str) -> Proyecto:
    """Carga un proyecto desde archivo Excel.

    Realiza estos pasos:
    1. Verifica que el archivo existe
    2. Lee las 5 hojas obligatorias (Proyecto, Unidades, Pilotes, Equipos, AsignacionEquipos)
    3. Valida los datos según MODELO_DATOS.md
    4. Construye el modelo interno

    Args:
        filepath: Ruta del archivo Excel.

    Returns:
        Objeto Proyecto completamente construido.

    Raises:
        FileNotFoundError: Si el archivo no existe.
        ValueError: Si hay errores de validación o estructura.
    """
    # Verificar que el archivo existe
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Archivo no encontrado: {filepath}")

    if not path.suffix.lower() in [".xlsx", ".xls"]:
        raise ValueError(f"Archivo debe ser Excel (.xlsx o .xls), recibido: {path.suffix}")

    # Abrir workbook
    try:
        wb = openpyxl.load_workbook(filepath)
    except Exception as e:
        raise ValueError(f"No se puede abrir archivo Excel: {str(e)}")

    # Leer 5 hojas obligatorias
    try:
        proyecto_data = _leer_hoja_proyecto(wb["Proyecto"])
        unidades_data = _leer_hoja_unidades(wb["Unidades"])
        pilotes_data = _leer_hoja_pilotes(wb["Pilotes"])
        equipos_data = _leer_hoja_equipos(wb["Equipos"])
        asignaciones_data = _leer_hoja_asignaciones(wb["AsignacionEquipos"])
    except KeyError as e:
        raise ValueError(f"Hoja faltante en Excel: {str(e)}")

    # Validar datos
    is_valid, errors = validate_project_data(
        proyecto_data, unidades_data, pilotes_data, equipos_data, asignaciones_data
    )

    if not is_valid:
        error_msg = "Errores en validación de datos:\n" + "\n".join(f"  - {e}" for e in errors)
        raise ValueError(error_msg)

    # Construir modelo
    proyecto = _construir_proyecto(
        proyecto_data, unidades_data, pilotes_data, equipos_data, asignaciones_data
    )

    return proyecto


def _leer_hoja_proyecto(ws: Worksheet) -> Dict[str, Any]:
    """Lee la hoja Proyecto.

    Estructura esperada:
        Proyecto | FechaInicio | RadioCritico_m | TiempoRestriccion_h
        [valor]  | [fecha]     | [número]       | [número]

    Returns:
        Diccionario con datos del proyecto.
    """
    data: Dict[str, Any] = {}

    for row in ws.iter_rows(min_row=1, max_row=2, values_only=True):
        if row[0] and row[0] == "Proyecto":
            # Primera fila contiene headers
            headers = row
            continue

    # Segunda fila contiene datos
    for row in ws.iter_rows(min_row=2, max_row=2, values_only=False):
        for cell in row:
            if cell and cell.value is not None:
                data[cell.parent.cell(row=1, column=cell.column).value] = cell.value

    return data


def _leer_hoja_unidades(ws: Worksheet) -> List[Dict[str, Any]]:
    """Lee la hoja Unidades.

    Estructura esperada:
        UnidadID | Nombre
        [id]     | [texto]
        ...

    Returns:
        Lista de diccionarios con datos de unidades.
    """
    data: List[Dict[str, Any]] = []
    headers: List[str] = []

    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx == 0:
            headers = [h for h in row if h]
            continue

        if not row or all(v is None for v in row):
            continue

        record = {}
        for col_idx, header in enumerate(headers):
            if col_idx < len(row):
                record[header] = row[col_idx]

        if record:
            data.append(record)

    return data


def _leer_hoja_pilotes(ws: Worksheet) -> List[Dict[str, Any]]:
    """Lee la hoja Pilotes.

    Estructura esperada:
        PiloteID | UnidadID | X      | Y
        [id]     | [id]     | [num]  | [num]
        ...

    Returns:
        Lista de diccionarios con datos de pilotes.
    """
    data: List[Dict[str, Any]] = []
    headers: List[str] = []

    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx == 0:
            headers = [h for h in row if h]
            continue

        if not row or all(v is None for v in row):
            continue

        record = {}
        for col_idx, header in enumerate(headers):
            if col_idx < len(row):
                record[header] = row[col_idx]

        if record:
            data.append(record)

    return data


def _leer_hoja_equipos(ws: Worksheet) -> List[Dict[str, Any]]:
    """Lee la hoja Equipos.

    Estructura esperada:
        EquipoID | Nombre | RendimientoPilotesDia | ModoInicio | PiloteInicio | ModoFin | PiloteFin
        [id]     | [text] | [int]                 | [AUTO/MAN] | [id/null]    | [A/M]   | [id/null]
        ...

    Returns:
        Lista de diccionarios con datos de equipos.
    """
    data: List[Dict[str, Any]] = []
    headers: List[str] = []

    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx == 0:
            headers = [h for h in row if h]
            continue

        if not row or all(v is None for v in row):
            continue

        record = {}
        for col_idx, header in enumerate(headers):
            if col_idx < len(row):
                record[header] = row[col_idx]

        if record:
            data.append(record)

    return data


def _leer_hoja_asignaciones(ws: Worksheet) -> List[Dict[str, Any]]:
    """Lee la hoja AsignacionEquipos.

    Estructura esperada:
        EquipoID | UnidadID | Prioridad
        [id]     | [id]     | [int]
        ...

    Returns:
        Lista de diccionarios con datos de asignaciones.
    """
    data: List[Dict[str, Any]] = []
    headers: List[str] = []

    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx == 0:
            headers = [h for h in row if h]
            continue

        if not row or all(v is None for v in row):
            continue

        record = {}
        for col_idx, header in enumerate(headers):
            if col_idx < len(row):
                record[header] = row[col_idx]

        if record:
            data.append(record)

    return data


def _construir_proyecto(
    proyecto_data: Dict[str, Any],
    unidades_data: List[Dict[str, Any]],
    pilotes_data: List[Dict[str, Any]],
    equipos_data: List[Dict[str, Any]],
    asignaciones_data: List[Dict[str, Any]],
) -> Proyecto:
    """Construye el modelo interno del proyecto.

    Args:
        proyecto_data: Datos validados del proyecto.
        unidades_data: Datos validados de unidades.
        pilotes_data: Datos validados de pilotes.
        equipos_data: Datos validados de equipos.
        asignaciones_data: Datos validados de asignaciones.

    Returns:
        Objeto Proyecto completamente construido.

    Raises:
        ValueError: Si hay error durante la construcción.
    """
    # Crear proyecto
    proyecto = Proyecto(
        nombre=str(proyecto_data["Proyecto"]).strip(),
        fecha_inicio=proyecto_data["FechaInicio"],
        radio_critico_m=float(proyecto_data["RadioCritico_m"]),
        tiempo_restriccion_h=float(proyecto_data["TiempoRestriccion_h"]),
    )

    # Crear unidades
    unidades_dict: Dict[str, UnidadEstructural] = {}
    for unidad_data in unidades_data:
        unidad = UnidadEstructural(
            id=str(unidad_data["UnidadID"]).strip(),
            nombre=str(unidad_data["Nombre"]).strip(),
        )
        unidades_dict[unidad.id] = unidad
        proyecto.agregar_unidad(unidad)

    # Crear pilotes y asignarlos a unidades
    for pilote_data in pilotes_data:
        pilote = Pilote(
            id=str(pilote_data["PiloteID"]).strip(),
            unidad_id=str(pilote_data["UnidadID"]).strip(),
            x=float(pilote_data["X"]),
            y=float(pilote_data["Y"]),
        )
        unidades_dict[pilote.unidad_id].agregar_pilote(pilote)

    # Crear equipos
    equipos_dict: Dict[str, Equipo] = {}
    for equipo_data in equipos_data:
        modo_inicio = ModoInicio[str(equipo_data["ModoInicio"]).upper()]
        modo_fin = ModoFin[str(equipo_data["ModoFin"]).upper()]

        equipo = Equipo(
            id=str(equipo_data["EquipoID"]).strip(),
            nombre=str(equipo_data["Nombre"]).strip(),
            rendimiento_pilotes_dia=int(equipo_data["RendimientoPilotesDia"]),
            modo_inicio=modo_inicio,
            pilote_inicio=str(equipo_data.get("PiloteInicio", "")).strip() or None,
            modo_fin=modo_fin,
            pilote_fin=str(equipo_data.get("PiloteFin", "")).strip() or None,
        )
        equipos_dict[equipo.id] = equipo
        proyecto.agregar_equipo(equipo)

    # Crear asignaciones
    for asig_data in asignaciones_data:
        asignacion = AsignacionEquipo(
            equipo_id=str(asig_data["EquipoID"]).strip(),
            unidad_id=str(asig_data["UnidadID"]).strip(),
            prioridad=int(asig_data["Prioridad"]),
        )
        proyecto.agregar_asignacion(asignacion)

    return proyecto
