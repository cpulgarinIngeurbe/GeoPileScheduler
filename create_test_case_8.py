"""Script para crear archivo Excel de prueba CASO 8 - Radio crítico pequeño."""

import openpyxl
from datetime import datetime


def crear_caso_prueba_8():
    """Crea CASO 8: Radio pequeño, muchos bloqueos."""
    wb = openpyxl.Workbook()

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Proyecto
    ws = wb.create_sheet("Proyecto", 0)
    ws["A1"], ws["B1"], ws["C1"], ws["D1"] = "Proyecto", "FechaInicio", "RadioCritico_m", "TiempoRestriccion_h"
    ws["A2"] = "Caso 8 - Radio Pequeño"
    ws["B2"] = datetime(2026, 1, 1)
    ws["C2"] = 5.0  # Radio pequeño → más bloqueos
    ws["D2"] = 24.0
    ws["B2"].number_format = "YYYY-MM-DD"

    # Unidades
    ws = wb.create_sheet("Unidades", 1)
    ws["A1"], ws["B1"] = "UnidadID", "Nombre"
    ws["A2"], ws["B2"] = "TORRE_A", "Torre A"
    ws["A3"], ws["B3"] = "TORRE_B", "Torre B"

    # Pilotes: 10 pilotes en grid (2x5)
    ws = wb.create_sheet("Pilotes", 2)
    ws["A1"], ws["B1"], ws["C1"], ws["D1"] = "PiloteID", "UnidadID", "X", "Y"

    pilotes_data = [
        ("P_A01", "TORRE_A", 0.0, 0.0), ("P_A02", "TORRE_A", 3.0, 0.0),
        ("P_A03", "TORRE_A", 6.0, 0.0), ("P_A04", "TORRE_A", 0.0, 3.0),
        ("P_A05", "TORRE_A", 3.0, 3.0),
        ("P_B01", "TORRE_B", 0.0, 0.0), ("P_B02", "TORRE_B", 3.0, 0.0),
        ("P_B03", "TORRE_B", 6.0, 0.0), ("P_B04", "TORRE_B", 0.0, 3.0),
        ("P_B05", "TORRE_B", 3.0, 3.0),
    ]

    for idx, (pid, uid, x, y) in enumerate(pilotes_data, start=2):
        ws[f"A{idx}"] = pid
        ws[f"B{idx}"] = uid
        ws[f"C{idx}"] = x
        ws[f"D{idx}"] = y

    # Equipos: 3 equipos
    ws = wb.create_sheet("Equipos", 3)
    ws["A1:G1"] = ["EquipoID", "Nombre", "RendimientoPilotesDia", "ModoInicio", "PiloteInicio", "ModoFin", "PiloteFin"]

    equipos_data = [
        ("EQ_A", "Equipo A", 2, "MANUAL", "P_A01", "AUTO", ""),
        ("EQ_B", "Equipo B", 2, "MANUAL", "P_A03", "AUTO", ""),
        ("EQ_C", "Equipo C", 1, "AUTO", "", "AUTO", ""),
    ]

    for idx, (eid, nombre, rend, m_ini, p_ini, m_fin, p_fin) in enumerate(equipos_data, start=2):
        ws[f"A{idx}"] = eid
        ws[f"B{idx}"] = nombre
        ws[f"C{idx}"] = rend
        ws[f"D{idx}"] = m_ini
        ws[f"E{idx}"] = p_ini
        ws[f"F{idx}"] = m_fin
        ws[f"G{idx}"] = p_fin

    # Asignaciones
    ws = wb.create_sheet("AsignacionEquipos", 4)
    ws["A1:C1"] = ["EquipoID", "UnidadID", "Prioridad"]

    asignaciones = [
        ("EQ_A", "TORRE_A", 1),
        ("EQ_B", "TORRE_A", 2),
        ("EQ_C", "TORRE_B", 1),
    ]

    for idx, (eid, uid, prio) in enumerate(asignaciones, start=2):
        ws[f"A{idx}"] = eid
        ws[f"B{idx}"] = uid
        ws[f"C{idx}"] = prio

    # Ancho columnas
    for ws in [wb.worksheets[i] for i in range(len(wb.sheetnames))]:
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20

    output_path = "data/input/caso_prueba_8.xlsx"
    wb.save(output_path)
    print(f"✅ Archivo creado: {output_path}")


if __name__ == "__main__":
    crear_caso_prueba_8()
