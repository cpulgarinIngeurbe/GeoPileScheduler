"""Script para crear archivo Excel de prueba CASO 5 - Simulación con paralelismo."""

import openpyxl
from datetime import datetime


def crear_caso_prueba_5():
    """Crea CASO 5: 1 unidad, 3 equipos, 10 pilotes (paralelismo + restricciones)."""
    wb = openpyxl.Workbook()

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Proyecto
    ws = wb.create_sheet("Proyecto", 0)
    ws["A1:D1"] = ["Proyecto", "FechaInicio", "RadioCritico_m", "TiempoRestriccion_h"]
    ws["A2"] = "Caso 5 - Simulación Paralela"
    ws["B2"] = datetime(2026, 1, 1)
    ws["C2"] = 8.0  # Radio medio → bloqueos moderados
    ws["D2"] = 24.0
    ws["B2"].number_format = "YYYY-MM-DD"

    # Unidades
    ws = wb.create_sheet("Unidades", 1)
    ws["A1:B1"] = ["UnidadID", "Nombre"]
    ws["A2:B2"] = ["TORRE_1", "Torre 1"]

    # Pilotes: 10 pilotes en grid (grid 4x3 con espaciado variable)
    ws = wb.create_sheet("Pilotes", 2)
    ws["A1:D1"] = ["PiloteID", "UnidadID", "X", "Y"]

    pilotes_data = [
        ("P_001", "TORRE_1", 0.0, 0.0),
        ("P_002", "TORRE_1", 6.0, 0.0),
        ("P_003", "TORRE_1", 12.0, 0.0),
        ("P_004", "TORRE_1", 18.0, 0.0),
        ("P_005", "TORRE_1", 0.0, 6.0),
        ("P_006", "TORRE_1", 6.0, 6.0),
        ("P_007", "TORRE_1", 12.0, 6.0),
        ("P_008", "TORRE_1", 18.0, 6.0),
        ("P_009", "TORRE_1", 0.0, 12.0),
        ("P_010", "TORRE_1", 6.0, 12.0),
    ]

    for idx, (pid, uid, x, y) in enumerate(pilotes_data, start=2):
        ws[f"A{idx}"] = pid
        ws[f"B{idx}"] = uid
        ws[f"C{idx}"] = x
        ws[f"D{idx}"] = y

    # Equipos: 3 equipos con distinto rendimiento
    ws = wb.create_sheet("Equipos", 3)
    ws["A1:G1"] = ["EquipoID", "Nombre", "RendimientoPilotesDia", "ModoInicio", "PiloteInicio", "ModoFin", "PiloteFin"]

    equipos_data = [
        ("EQ_A", "Equipo Rápido", 3, "AUTO", "", "AUTO", ""),
        ("EQ_B", "Equipo Medio", 2, "AUTO", "", "AUTO", ""),
        ("EQ_C", "Equipo Lento", 1, "AUTO", "", "AUTO", ""),
    ]

    for idx, (eid, nombre, rend, m_ini, p_ini, m_fin, p_fin) in enumerate(equipos_data, start=2):
        ws[f"A{idx}"] = eid
        ws[f"B{idx}"] = nombre
        ws[f"C{idx}"] = rend
        ws[f"D{idx}"] = m_ini
        ws[f"E{idx}"] = p_ini
        ws[f"F{idx}"] = m_fin
        ws[f"G{idx}"] = p_fin

    # Asignaciones
    ws = wb.create_sheet("AsignacionEquipos", 4)
    ws["A1:C1"] = ["EquipoID", "UnidadID", "Prioridad"]

    asignaciones = [
        ("EQ_A", "TORRE_1", 1),
        ("EQ_B", "TORRE_1", 2),
        ("EQ_C", "TORRE_1", 3),
    ]

    for idx, (eid, uid, prio) in enumerate(asignaciones, start=2):
        ws[f"A{idx}"] = eid
        ws[f"B{idx}"] = uid
        ws[f"C{idx}"] = prio

    # Ancho columnas
    for sheet in wb.worksheets:
        for col in sheet.columns:
            sheet.column_dimensions[col[0].column_letter].width = 20

    output_path = "data/input/caso_prueba_5.xlsx"
    wb.save(output_path)
    print(f"✅ Archivo creado: {output_path}")


if __name__ == "__main__":
    crear_caso_prueba_5()
